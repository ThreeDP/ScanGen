from createTable import create_table
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Rotulo.xlsx')
ws = wb['Envolope']
count = 0

table = create_table()

def te(n, table):
    positions = ['C5', 'H5', 'C12', 'H12', 'C19', 'H19', 'C26', 'H26', 'C33', 'H33']
    for p in positions:
        try:
            ws[p] = table[n]['Nome']
            print(p, n)
        except Exception as e:
            print(e)
            ws[p] = ""
            
        n+= 1
    
    string = './patterns/Clientes' + str(n) + '.xlsx'
    wb.save(string)
    return n

while count < len(table):
    count = te(count, table)
    print(count)
    