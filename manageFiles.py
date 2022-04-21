from catchCsv import catch_csv
from iosActions import find_files, make_dir
from openpyxl import load_workbook

wb = load_workbook('Rotulo.xlsx')
ws = wb['Envolope']

def create_table(path):
    
    files = find_files(path)
    table = list()

    for file in files:
        pathfile = path + file
        data = catch_csv(pathfile)

        y = data[0].replace('"', '') # Elimina as aspas da string (Cabeçalho)
        header = y.split(';') # Divide a string em um array (Cabeçalho)
        for line in range(1, len(data), 1):
            x = data[line].replace('"', '')
            words = x.split(';')
            
            row = dict().fromkeys(header, '') # Cria um dicionario vazio.

            for column_index in range(0, len(header) - 1, 1): # Pula o cabeçalho e realiza o laço de inserção de dados.
                    row[header[column_index]] = words[column_index] # Preenche o dicionário criado com os dados

            for i in row.values(): # Valida se a row está vazia antes de inserir.
                if i != '':
                    table.append(row)
                    break
        
    return table

def insert_data(name_index, table, path):
    positions = ['C5', 'H5', 'C12', 'H12', 'C19', 'H19', 'C26', 'H26', 'C33', 'H33'] # Posições a serem inserido os dados.

    for position in positions:
        try:
            ws[position] = table[name_index]['Nome'] # Inserção do Nome do cliente na posição da interação.

        except IndexError as index: # Limpa os nomes duplicados da interação anterior caso houver estouro da lista table.
            ws[position] = "" 
           

        except Exception as e:
            print(e)
            
        name_index+= 1
    
    label_file = path + 'Clientes' + str(name_index) + '.xlsx' 
    wb.save(label_file) # Salva o arquivo na pasta e com o nome estabelecido na string label_file.
    return name_index

def import_files(path): # Necessário o envio de cada arquivo com uma key diferente.
    files = request.files
    for file_index in range(0, len(files), 1):
        file = request.files.get(str(file_index))
        file.save(path + file.filename)